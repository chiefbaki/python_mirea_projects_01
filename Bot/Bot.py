# Встроенные библиотеки
import re
import time
import datetime
from typing import Any

import matplotlib.pyplot as plt
import openpyxl
import requests

# Сторонние библиотеки
import vk_api
from config import token
from bs4 import BeautifulSoup
from vk_api import VkUpload
from vk_api.keyboard import VkKeyboard, VkKeyboardColor
from vk_api.longpoll import VkLongPoll, VkEventType
from vk_api.utils import get_random_id
from PIL import Image, ImageDraw, ImageFont
from urllib.request import urlopen

# Самописные модули
from five_day_weather import TodayWeather
from helper import *
from database import *
from weather import *


class VkBot:
    def __init__(self) -> None:
        """
        Конструктор
        """
        self.vk_session = vk_api.VkApi(token=token)
        self.vk = self.vk_session.get_api()
        self.longpoll = VkLongPoll(self.vk_session)
        self.users_to_set_group: set = set()
        self.users_to_set_teacher: set = set()
        self.users_to_get_teacher: list = []
        self.last_schedule_file_update: time
        self.schedule_data: list
        if True:
            self._update_schedule_file()
        else:
            self._parse_schedule_file()
        Debug('Bot init', key='SRT')

    def start_listen(self) -> None:
        """
        Слушатель событий вк
        """
        Debug('Start listen', key='SRT')
        for event in self.longpoll.listen():
            if event.type == VkEventType.MESSAGE_NEW and event.text and event.to_me and event.from_user:
                Debug('from {} text = \'{}\''.format(event.user_id, event.text), key='MSG')
                self._command_handler(event.user_id, event.text.lower())

    def _command_handler(self, user_id: int, text: str) -> None:
        """
        Обработчик команд
        """
        for i in range(len(self.users_to_get_teacher)):
            if self.users_to_get_teacher[i][0] == user_id:
                match text:
                    case 'на сегодня':
                        self._show_today_teacher_schedule(user_id, self.users_to_get_teacher[i][1])
                    case 'на завтра':
                        self._show_today_teacher_schedule(user_id, self.users_to_get_teacher[i][1], 1)
                    case 'на эту неделю':
                        self._show_teacher_week_schedule(user_id, self.users_to_get_teacher[i][1])
                    case 'на следующую неделю':
                        self._show_teacher_week_schedule(user_id, self.users_to_get_teacher[i][1], 1)
                del self.users_to_get_teacher[i]
                return
        match text:
            case "начать":
                user_data = self.vk.users.get(user_id=user_id)[0]
                self._send_message(user_id, 'Привет {}!\nЯ бот, который показывает расписание института информационных технологий РТУ МИРЭА, ' \
             'а так же я умею показывать погоду и ситуацию по коронавирусу.\n\n' \
             'Для начала введи свою группу. \n' \
             'Набери "помощь", чтобы посмотреть мои возможности. Но только после ввода группы!'.format(user_data['first_name']))
                Debug(f'Start new user: {user_data["first_name"]} {user_data["last_name"]}')
                self._add_user_to_edit_group_list(user_id)
                return
            case 'бот':
                self._show_schedule_keyboard(user_id)
                return
            case 'на сегодня':
                self._show_today_schedule(user_id)
                return
            case 'на завтра':
                self._show_tomorrow_schedule(user_id)
                return
            case 'на эту неделю':
                self._show_week_schedule(user_id)
                return
            case 'на следующую неделю':
                self._show_week_schedule(user_id, week_delta=1)
                return
            case 'неделя?':
                self._show_current_week(user_id)
                return
            case 'группа?':
                self._show_user_group(user_id)
                return
            case 'помощь':
                self._show_help_message(user_id)
                return
            case 'корона':
                self._show_corona_all_stat(user_id)
                return
            case 'погода':
                self._show_weather_keyboard(user_id)
                return
            case 'сейчас':
                self._show_now_weather(user_id)
                return
            case 'на 5 дней':
                self._show_5days_weather(user_id)
                return
            case 'сегодня':
                self._show_today_weather(user_id)
                return
            case 'завтра':
                self._show_tomorrow_weather(user_id)
                return


        combo_cmd = text.split(' ')
        match combo_cmd[0]:
            case 'бот':
                if len(combo_cmd) == 2:
                    if combo_cmd[1].lower() in ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье']:
                        self._show_week_day_schedule(user_id, combo_cmd[1].lower())
                    else:
                        self._edit_user_group(user_id, combo_cmd[1])
                    return
                elif len(combo_cmd) == 3:
                    if combo_cmd[1].lower() in ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье']:
                        if self._validate_group_slug(combo_cmd[2]):
                            self._edit_user_group(user_id, combo_cmd[2])
                            self._show_week_day_schedule(user_id, combo_cmd[1].lower())
                        else:
                            self._send_message(user_id, 'Неверный формат или группа не найдена!\n\nФормат: \'АБВГ-12-34\'')
                        return
            case 'найти':
                self._show_teacher_keyboard(user_id, combo_cmd[1:])
                return
            case 'корона':
                self._show_corona_local_data(user_id, combo_cmd[1:])
                return

        if str(user_id) in self.users_to_set_group:
            self._edit_user_group(user_id, text)
            return

        if str(user_id) in self.users_to_set_teacher:
            self._show_teacher_period_keyboard(user_id, text)
            return

        Debug(f'Command {text} not found for id: {user_id}', key='CNF')
        self._send_message(user_id, 'Неизвестная команда\nЧто бы получить список команд напиши \'{}\''.format('помощь'.title()))

    def _get_week_schedule(self, group: str, date: datetime.datetime, with_reformat: bool = True) -> list:
        """
        Возвращает расписанию на неделю, дату которой передали
        """
        now = date.isocalendar()
        week = now.week - 5
        week_even = (week + 1) % 2  # Является ли неделя чётной
        column = self._get_group_column(group)
        out = []
        tmp = []
        for i in range(2 + week_even, len(self.schedule_data[column]), 2):  # Каждый второй, со смещением по недели
            tmp.append(
                [self.schedule_data[column][i],  # Предмет
                 self.schedule_data[column + 1][i],  # Вид занятий
                 self.schedule_data[column + 2][i],  # Преподаватель
                 self.schedule_data[column + 3][i]]  # Кабинет
            )
            if len(tmp) == 6:
                out.append(tmp)
                tmp = []

        for i in range(len(out)):
            for j in range(6):
                out[i][j][0] = self._reformat_subject_name(out[i][j][0], week_number=week,
                                                           ignore_weeks=(not with_reformat))
                out[i][j][1] = self._reformat_double_pair(out[i][j][1])
                out[i][j][2] = self._reformat_double_pair(out[i][j][2])
                out[i][j][3] = self._reformat_double_pair(out[i][j][3])
        Debug(f'Getting {group} week schedule', key='GET')
        return out

    def _get_day_schedule(self, group: str, date: datetime.datetime) -> list:
        """
        Возвращает массив с расписанием на текущий день
        """
        week = self._get_week_schedule(group, date)
        week_index = date.isocalendar().weekday - 1
        if week_index == 6:
            return [[] * 4] * 6
        Debug(f'Getting {group} day schedule', key='GET')
        return week[week_index]

    def _get_user_group(self, user_id: int) -> str or None:
        """
        Получает группу пользователя или ошибка
        """
        group = Database().fetch_one(table='groups', condition=f'user_id = {user_id}')
        if group:
            Debug(f'Find {user_id} group: {group}', key='FND')
            return group[1]
        else:
            self._send_message(user_id, 'Группа не выбрана, для выбора группы, напишите \n\'{}\' и номер группы'.format('бот'.title()))
        return None

    def _get_current_week(self) -> int:
        """
        Возвращает номер текущей недели
        """
        Debug(f'Getting current week', key='GET')
        return datetime.datetime.now().isocalendar().week - 5

    def _get_group_column(self, group) -> int or None:
        """
        Ищет столбец группы в расписании
        """
        for i in range(0, len(self.schedule_data), 4):
            if self.schedule_data[i][0] == group:
                Debug(f'Find in file group: {group} column: {i + 1}', key='FND')
                return i
        return None

    def _get_string_date(self, date: datetime.datetime, with_week_day: bool = False) -> str:
        """
        Преобразует дату в строку с датой
        """
        result = ''
        if with_week_day:
            result += ['понедельник', 'вторник', 'среду', 'четверг', 'пятницу', 'субботу', 'воскресенье'][date.isocalendar().weekday - 1] + " "
        result += str(date.day) + " " + ['января', 'февраля', 'марта', 'апреля', 'мая', 'июня', 'июля', 'августа', 'сентября', 'октября',
                'ноября', 'декабря'][date.month % 12 - 1]
        return result

    def _get_teacher_full_name(self, teacher: str) -> set[str]:
        """
        Получение полного имени преподавателей из расписания
        """
        result = set()
        for i in range(2, len(self.schedule_data), 4):
            for j in range(2, len(self.schedule_data[i])):
                tmp = self.schedule_data[i][j].split('\n')
                if len(tmp) > 0:
                    if tmp[0].split(' ')[0] == teacher:
                        result.add(tmp[0] if tmp[0][-1] == '.' else tmp[0] + '.')  # Исправление косяков расписания
                    elif tmp[-1].split(' ')[0] == teacher:

                        result.add(tmp[-1] if tmp[-1][-1] == '.' else tmp[-1] + '.')
        Debug(f'Getting {teacher} fullname', key='GET')
        return result

    def _get_teacher_week_schedule(self, teacher: str, date: datetime.datetime, with_reformat: bool = True) -> list:
        """
        Возвращает расписание преподавателя на указанную неделю
        """
        now = date.isocalendar()
        week = now.week - 5
        week_even = (week + 1) % 2  # Является ли неделя чётной
        out = []
        tmp = []

        for j in range(2 + week_even, len(self.schedule_data[0]), 2):
            para = []  # одна пара
            for i in range(2, len(self.schedule_data), 4):  # Слева на права
                tmp_teachers = self.schedule_data[i][j].split('\n')  # для сдвоенных пар
                if len(tmp_teachers) > 0:
                    t1 = tmp_teachers[0] if tmp_teachers[0][-1] == '.' else tmp_teachers[0] + '.'
                    t2 = tmp_teachers[-1] if tmp_teachers[-1][-1] == '.' else tmp_teachers[-1] + '.'
                    if t1 == teacher:
                        para = [
                            self.schedule_data[i - 2][j].split('\n')[0],  # Предмет
                            self.schedule_data[i - 1][j].split('\n')[0],  # Вид
                            self.schedule_data[i - 2][0],  # Группа
                            self.schedule_data[i + 1][j].split('\n')[0]  # Аудитория
                        ]
                        break
                    elif t2 == teacher:
                        para = [
                            self.schedule_data[i - 2][j].split('\n')[-1],  # Предмет
                            self.schedule_data[i - 1][j].split('\n')[-1],  # Вид
                            self.schedule_data[i - 2][0],  # Группа
                            self.schedule_data[i + 1][j].split('\n')[-1]  # Аудитория
                        ]
                        break
                    # Останавливаем смешение вправо, если нашли
            tmp.append(para)  # Добавляем пару, даже если она пустая
            if (j - week_even) % 12 == 0:
                out.append(tmp)
                tmp = []

        for i in range(len(out)):
            for j in range(6):
                if len(out[i][j]) > 1:
                    out[i][j][0] = self._reformat_subject_name(out[i][j][0], week_number=week,
                                                               ignore_weeks=(not with_reformat))
        return out

    def _get_day_teacher_schedule(self, teacher: str, date: datetime.datetime) -> list:
        """
        Возвращает расписание преподавателя на переданный день
        """
        week = self._get_teacher_week_schedule(teacher, date)
        week_index = date.isocalendar().weekday - 1
        if week_index == 6:
            return [[] * 4] * 6
        return week[week_index]

    def _get_corona_stat(self, extra_url: str = '') -> tuple[str, list[Any], list[Any]]:
        """
        Возвращает статистику коронавируса на сегодня определённой области
        """
        page = requests.get('https://coronavirusstat.ru' + extra_url)  # Получаем страницу
        soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
        result = soup.find(string='Прогноз заражения на 10 дней').find_parent('div', {
            'class': 'border rounded mt-3 mb-3 p-3'})
        status = result.find('h6', 'text-muted').getText()[:-17]
        data = result.findAll('div', {'class': 'col col-6 col-md-3 pt-4'})
        plus = [] * 4
        value = [] * 4
        for i in range(4):
            value.append(data[i].find('div', 'h2').getText())
            plus.append(data[i].find('span', {'class': 'font-weight-bold'}).getText())
        return status, value, plus

    def _get_corona_all_stat(self) -> tuple[list[Any], list[float], list[float], list[float], list[float]]:
        """
        Возвращает статистику коронавируса на последние 10 дней
        """
        page = requests.get('https://coronavirusstat.ru' + '/country/russia')  # Получаем страницу
        soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
        result = soup.find('table', {'class': 'table table-bordered small'}).findAll('tr')
        days = []
        active = []
        cured = []
        died = []
        cases = []
        stats = []
        ml = 1000000

        for i in range(1, 11):
            days.append(result[i].find('th').getText())
            for a in result[i].findAll('td'):
                stats.append(int(a.getText().split(' ')[1]))
        for i in range(0, len(stats), 4):
            active.append(stats[i] / ml)
        for i in range(1, len(stats), 4):
            cured.append(stats[i] / ml)
        for i in range(2, len(stats), 4):
            died.append(stats[i] / ml)
        for i in range(3, len(stats), 4):
            cases.append(stats[i] / ml)

        days = list(reversed(days))
        active = list(reversed(active))
        cured = list(reversed(cured))
        died = list(reversed(died))
        cases = list(reversed(cases))

        return days, active, cured, died, cases

    def _get_wind_slug(self, speed: float) -> str:
        """
        Возвращает словесное название  ветра по шкале Бофорта
        """
        s = 'Ураган'
        if speed <= 0.2:
            s = 'Штиль'
        elif speed <= 1.5:
            s = 'Тихий'
        elif speed <= 3.3:
            s = 'Лёгкий'
        elif speed <= 5.4:
            s = 'Слабый'
        elif speed <= 7.9:
            s = 'Умеренный'
        elif speed <= 10.7:
            s = 'Свежий'
        elif speed <= 13.8:
            s = 'Сильный'
        elif speed <= 17.1:
            s = 'Крепкий'
        elif speed <= 20.7:
            s = 'Очень крепкий'
        elif speed <= 24.4:
            s = 'Шторм'
        elif speed <= 28.4:
            s = 'Сильный шторм'
        elif speed <= 32.6:
            s = 'Жестокий шторм'
        return s

    def _get_wind_deg_slug(self, deg: int) -> str:
        """
        Возращает буквенное направление ветра
        """
        s = 'северный'
        if deg <= 22.5:
            s = 'северный'
        elif deg <= 67.5:
            s = 'северо-восточный'
        elif deg <= 112.5:
            s = 'восточный'
        elif deg <= 157.5:
            s = 'юго-восточный'
        elif deg <= 202.5:
            s = 'южный'
        elif deg <= 247.5:
            s = 'юго-западный'
        elif deg <= 292.5:
            s = 'западный'
        elif deg <= 292.5:
            s = 'северо-западный'
        return s


    def _show_now_weather(self, user_id: int) -> None:
        """
        Показывает текущую погоду в Москве
        """
        upload = VkUpload(self.vk_session)
        attachments = []
        image = Image.open('weather_pattern.jpg')
        # img = image.resize((400, 500))
        img = image.resize((400, 260))
        img = img.convert('RGB')
        idraw = ImageDraw.Draw(img)
        title = ImageFont.truetype(font='lato.ttf', size=30)
        font = ImageFont.truetype(font='lato.ttf', size=18)
        # title2 = ImageFont.truetype(size=30)
        idraw.text((10, 10), 'Погода в Москве', font=title, fill="white")
        response = requests.get("http://api.openweathermap.org/data/2.5/weather",
                                params={'q': 'moscow,ru', 'units': 'metric', 'APPID': '3051c8da538a2e104685783dae4796f4', 'lang': 'ru'})
        info = response.json()
        img_name = info['weather'][0]['icon']
        im = Image.open(urlopen('https://openweathermap.org/img/wn/{}@4x.png'.format(img_name)))
        img.paste(im, (95, 20), im.convert('RGBA'))
        status = info['weather'][0]['description'].capitalize()
        temp = str(int(info['main']['temp_min'])) + ' - ' + str(int(info['main']['temp_max']))
        pressure = str(int(float(info['main']['pressure']) / 1.33))
        humidity = str(info['main']['humidity'])
        wind_speed = info['wind']['speed']
        wind_slug = self._get_wind_slug(float(wind_speed)).lower()
        wind_deg_slug = self._get_wind_deg_slug(info['wind']['speed'])
        weather = '{}, температура: {}°C\nДавление: {} мм рт. сб., влажность: {}%\nВетер: {}, {} м/с, {}'.format(status, temp, pressure, humidity, wind_slug, wind_speed,
                                                  wind_deg_slug)
        idraw.text((10, 185), weather, font=font, fill="white")
        img.save('data/weather_card.png')

        photo = upload.photo_messages('data/weather_card.png')[0]
        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

        self._send_message_with_attachments(user_id=user_id, text='Сейчас', attachments=attachments)


    def _show_tomorrow_weather(self, user_id: int) -> None:
        upload = VkUpload(self.vk_session)
        self._send_message(user_id, 'Погода в Москве на завтра')
        attachments = []
        tw = TomorrowWeather()
        tw.show_pic
        photo = upload.photo_messages('data/weather_tom.png')[0]
        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

        self._send_message_with_attachments(user_id=user_id, attachments=attachments)
        self._send_message(user_id, tw.get_weather_tomorrow)

    def _show_5days_weather(self, user_id: int) -> None:
        upload = VkUpload(self.vk_session)
        self._send_message(user_id, 'Погода в Москве на ближайшие 5 дней')
        attachments = []
        fd = Weather5Day()
        fd.show_pic
        photo = upload.photo_messages('data/weather_5day.png')[0]
        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

        self._send_message_with_attachments(user_id=user_id, attachments=attachments)
        self._send_message(user_id, fd.get_info)

    def _show_today_weather(self, user_id: int) -> None:
        upload = VkUpload(self.vk_session)
        self._send_message(user_id, 'Погода в Москве на сегодня')
        attachments = []
        tw = TodayWeather()
        tw.show_pic
        photo = upload.photo_messages('data/weather_tom.png')[0]
        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))

        self._send_message_with_attachments(user_id=user_id, attachments=attachments)
        self._send_message(user_id, tw.get_weather_today)

    def _show_corona_all_stat(self, user_id: int) -> None:
        """
        Показывает статистику коронавируса на сегодня и выводит график
        """
        days, active, cured, died, cases = self._get_corona_all_stat()
        graf_data = {
            'Активных': active,
            'Вылечено': cured,
            'Умерло': died,
        }
        for i in range(len(days)):
            days[i] = days[i][:-5]
        fig, ax = plt.subplots()
        ax.stackplot(days, graf_data.values(),
                     labels=graf_data.keys(), alpha=0.8)
        ax.legend(loc='upper left')
        ax.set_title('Россия - Детальная статистика - коронавирус')
        ax.set_ylabel('Количество - Миллионы')
        fig.savefig('data\graf.png')
        upload = VkUpload(self.vk_session)
        attachments = []
        photo = upload.photo_messages('data/' + 'graf.png')[0]
        attachments.append("photo{}_{}".format(photo["owner_id"], photo["id"]))
        self._send_message_with_attachments(user_id=user_id,
                                            text=self._reformat_corona_data('Россия', self._get_corona_stat()),
                                            attachments=attachments)

    def _show_weather_keyboard(self, user_id: int) -> None:
        """
        Показывает клавиатуру для выбора периода погоды
        """
        keyboard = VkKeyboard(one_time=True)
        keyboard.add_button('сейчас', color=VkKeyboardColor.PRIMARY)
        keyboard.add_button('сегодня', color=VkKeyboardColor.POSITIVE)
        keyboard.add_button('завтра', color=VkKeyboardColor.POSITIVE)
        keyboard.add_line()
        keyboard.add_button('на 5 дней', color=VkKeyboardColor.POSITIVE)
        self._send_message(user_id, text='Показать погоду в Москве', custom_keyboard=keyboard)

    def _show_corona_local_data(self, user_id: int, region_list: list) -> None:
        """
        Показывает статистику коронавируса по региону
        """
        if len(region_list) > 0:
            region = region_list[0].title()
            page = requests.get('https://coronavirusstat.ru/country/russia')  # Получаем страницу
            soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
            result = soup.findAll('div', {'class': 'c_search_row'})
            d = ''
            rg = 'Не найден, поэтому Россия'
            for x in result:
                tmp = x.find('span', 'small').find('a')
                if region.title() in tmp.getText().split(' '):
                    rg = tmp.getText()
                    d = tmp.get('href')
                    break
            self._send_message(user_id, self._reformat_corona_data(rg, self._get_corona_stat(d)))

    def _show_today_teacher_schedule(self, user_id: int, teacher: str, day_delta: int = 0) -> None:
        """
        Выводит расписания на сегодня день(со смещением)
        """
        now = datetime.datetime.now() + datetime.timedelta(days=day_delta)
        teacher = self._reformat_teacher_name(teacher)
        if self._validate_teacher_name(teacher):
            schedule = self._get_day_teacher_schedule(teacher, now)
            self._send_message(user_id=user_id, text=self._reformat_day_schedule(schedule, now, teacher_header=teacher))

    def _show_teacher_week_schedule(self, user_id: int, teacher: str, week_delta: int = 0) -> None:
        """
        Выводит расписание преподавателя на неделю
        """
        now = datetime.datetime.now() + datetime.timedelta(weeks=week_delta)
        day_date = now - datetime.timedelta(days=now.isocalendar().weekday - 1)
        result = ''
        teacher = self._reformat_teacher_name(teacher)
        if self._validate_teacher_name(teacher):
            schedule = self._get_teacher_week_schedule(teacher, now)
            for i in range(6):
                result += self._reformat_day_schedule(schedule[i], date=day_date, teacher_header=teacher,
                                                      week_format=True)
                day_date += datetime.timedelta(days=1)
            self._send_message(user_id, result)

    def _show_teacher_period_keyboard(self, user_id: int, teacher: str) -> None:
        """
        Показывает клавиатуру для выбора периода расписания преподавателя
        """
        tmp = teacher.split(' ')
        if len(tmp) == 2:
            teacher = tmp[0].title() + ' ' + tmp[1].upper()
            if self._validate_teacher_name(teacher):
                Debug(f'Show period for {teacher}', key='SHW')
                # Создаём клавиатуру
                keyboard = VkKeyboard(one_time=True)
                keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
                keyboard.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
                keyboard.add_line()
                keyboard.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
                keyboard.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)

                self._clear_wait_lists(user_id)
                self._add_user_to_get_teacher_list(user_id, teacher)
                self._send_message(user_id, text='Показать расписание преподавателя {} ...'.format(teacher),
                                   custom_keyboard=keyboard)
                return
        self._send_message(user_id, text='Преподаватель не найден')

    def _show_teacher_keyboard(self, user_id: int, teacher: list = None):
        """
        Показывает клавиатуру выбора преподавателя, либо выбора периода, для показа расписания
        """
        name = ''
        if len(teacher) == 2:
            name = teacher[0].title() + ' ' + teacher[1].upper()
        elif len(teacher) == 1:  # Только фамилия
            tmp = []
            for a in self._get_teacher_full_name(teacher[0].title()):
                tmp.append(a)
            if len(tmp) == 1:  # Если 1, то
                name = tmp[0]
            elif len(tmp) > 1:
                self._add_user_to_set_teacher_list(user_id)  # Добавляем пользователя в список ожидания
                keyboard = VkKeyboard(one_time=True)
                for i in range(len(tmp)):
                    keyboard.add_button(tmp[i], color=VkKeyboardColor.SECONDARY)
                    if i % 2 and i != len(tmp) - 1:  # Каждый второй, но не последний
                        keyboard.add_line()
                self._send_message(user_id=user_id, text='Выберите преподавателя', custom_keyboard=keyboard)
                Debug(f'Show teacher keyboard id {user_id}')
                return
        if len(name) > 1:
            self._show_teacher_period_keyboard(user_id, name)
            return
        self._send_message(user_id, 'Преподаватель не найден')

    def _show_help_message(self, user_id: int) -> None:
        """
        Отправляет подсказку с командами
        """
        self._send_message(user_id, 'Список команд:\n\nНачать - Запускает бота\nБот - Показывает клавиатуру для выбора периода расписания\n' \
            'Бот <номер группы> - Запоминает группу и показывает клавиатуру\n' \
            'Бот <день недели> - Показывает расписание на  выбранный день\n' \
            'Бот <день недели> <номер группы> - Сохраняет группу и показывает расписание на день\n\n' \
            'Найти <фамилия преподавателя> [И.О.] - Получить расписание преподавателя за определённый период\n\n' \
            '<> - Обязательные аргументы\n[] - По желанию')

    def _show_week_day_schedule(self, user_id: int, day: str) -> None:
        """
        Показывает расписание на определённый день недели
        """
        group = self._get_user_group(user_id)
        if group:
            date = datetime.datetime.now()
            if self._get_current_week() % 2 == 0:  # Если она не чётная
                date -= datetime.timedelta(weeks=1)
            odd = self._get_week_schedule(group=group, date=date, with_reformat=False)
            date += datetime.timedelta(weeks=1)
            even = self._get_week_schedule(group=group, date=date, with_reformat=False)
            index = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье'].index(day)
            o = self._reformat_day_schedule(data=odd[index], with_header=False)  # Нечётный день
            e = self._reformat_day_schedule(data=even[index], with_header=False)  # Чётный день
            result = 'Расписание на {}, нечётной недели\n'.format(day.title()) + o + '\n\n' + 'Расписание на {} чётной недели\n'.format(day.title()) + e
            self._send_message(user_id, result)

    def _show_today_schedule(self, user_id: int) -> None:
        """
        Выводит расписание на сегодняшний день
        """
        now = datetime.datetime.now()
        group = self._get_user_group(user_id)
        if group:
            schedule = self._get_day_schedule(group, now)
            self._send_message(user_id=user_id, text=self._reformat_day_schedule(schedule, now))

    def _show_tomorrow_schedule(self, user_id: int) -> None:
        """
        Выводит расписание на завтрашний день
        """
        now = datetime.datetime.now() + datetime.timedelta(days=1)
        group = self._get_user_group(user_id)
        if group:
            schedule = self._get_day_schedule(group, now)
            Debug(f'Show {group} tomorrow schedule for id: {user_id}', key='SHW')
            self._send_message(user_id=user_id, text=self._reformat_day_schedule(schedule, now))

    def _show_week_schedule(self, user_id: int, week_delta: int = 0) -> None:
        """
        Выводит расписание на текущую неделю
        """
        now = datetime.datetime.now() + datetime.timedelta(weeks=week_delta)
        day_date = now - datetime.timedelta(days=now.isocalendar().weekday - 1)
        group = self._get_user_group(user_id)
        result = ''
        if group:
            schedule = self._get_week_schedule(group, now)
            for i in range(6):
                result += self._reformat_day_schedule(schedule[i], date=day_date, week_format=True)
                day_date += datetime.timedelta(days=1)
            Debug(f'Show {group} week schedule for id: {user_id}', key='SHW')
            self._send_message(user_id, result)

    def _show_current_week(self, user_id: int) -> None:
        """
        Выводит пользователю номер текущей недели
        """
        self._send_message(user_id, 'Идёт {} неделя'.format(self._get_current_week()))

    def _show_user_group(self, user_id: int) -> None:
        """
        Выводит пользователю номер выбранной группы или ошибку
        """
        group = self._get_user_group(user_id)
        if group:
            self._send_message(user_id, 'Я показываю расписание группы {}'.format(group))

    def _show_schedule_keyboard(self, user_id: int) -> None:
        """
        Показать клавиатуру выбора расписания
        """
        Debug(f'Show schedule keyboard for id: {user_id}')
        self._send_message(user_id, 'Показать расписание ...', keyboard=1)

    def _add_user_to_edit_group_list(self, user_id) -> None:
        """
        Добавляет пользователя в список обновления группы
        """
        self.users_to_set_group.add(str(user_id))
        Debug(f'User add to set group list, uid: {user_id}', key='SET')

    def _add_user_to_set_teacher_list(self, user_id) -> None:
        """
        Добавляет пользователя в список выбора преподавателя
        """
        self.users_to_set_teacher.add(str(user_id))
        Debug(f'User add to set teacher list, uid: {user_id}', key='SET')

    def _add_user_to_get_teacher_list(self, user_id, teacher: str) -> None:
        """
        Добавляет пользователя в список выбора преподавателя
        """
        self.users_to_get_teacher.append([user_id, teacher])
        Debug(f'User add to get teacher list, uid: {user_id}', key='SET')

    def _edit_user_group(self, user_id: int, group_slug: str) -> None:
        """
        Изменить группу пользователя или выдать ошибку
        """
        group_slug = group_slug.upper()
        if self._validate_group_slug(group_slug):
            db = Database()
            if not db.fetch_one(table='groups', condition=f'user_id = {user_id}'):
                db.insert_one(table='groups', data=[user_id, group_slug])
                Debug(f'Set user group: {group_slug} uid: {user_id}', key='SET')
            else:
                db.update_one(table='groups', sets=f'group_slug = \'{group_slug}\'',
                              condition=f'user_id = {user_id}')
                Debug(f'ReSet user group: {group_slug} uid: {user_id}', key='RST')
            self._send_message(user_id, 'Я запомнил, что ты учишься в группе {}'.format(group_slug), 1)
            self._clear_wait_lists(user_id)  # Убираем из списка ожидания
            self._show_schedule_keyboard(user_id=user_id)  # Показываем клавиатуры выбора
            del db
        else:
            self._send_message(user_id, 'Неверный формат или группа не найдена!\n\nФормат: \'АБВГ-12-34\'')
            Debug(f'Invalid group format {group_slug} uid: {user_id}', key='INV')

    def _clear_wait_lists(self, user_id: int) -> None:
        """
        Убирает пользователя из списка ожидания
        """
        self.users_to_set_group.discard(str(user_id))
        self.users_to_set_teacher.discard(str(user_id))
        for i in range(len(self.users_to_get_teacher)):
            if self.users_to_get_teacher[i][0] == user_id:
                del self.users_to_get_teacher[i]
                break

    def _validate_teacher_name(self, teacher: str) -> bool:
        """
        Проверяет наличие преподавателя в файлах расписания
        """
        for i in range(2, len(self.schedule_data), 4):
            for j in range(2, len(self.schedule_data[i])):
                tmp = self.schedule_data[i][j].split('\n')  # для сдвоенных пар
                if len(tmp) > 0:
                    # Исправляем отсутствие точки у некоторых преподавателей
                    t1 = tmp[0] if tmp[0][-1] == '.' else tmp[0] + '.'
                    t2 = tmp[-1] if tmp[-1][-1] == '.' else tmp[-1] + '.'
                    if t1 == teacher or t2 == teacher:
                        Debug(f'Find teacher {t1} or {t2}', key='FND')
                        return True
        return False

    def _validate_group_slug(self, group_slug: str) -> bool:
        """
        Проверка на валидность номера группы по маске и списку групп
        """
        group_slug = group_slug.upper()
        if re.match(r'\w{4}-\d{2}-\d{2}', group_slug):
            if self._get_group_column(group_slug):
                return True
        return False

    def _reformat_corona_data(self, region: str, data: tuple) -> str:
        """
        Реформат данных коронавируса на 1 день
        """
        status, value, plus = data
        return '{}\n\nРегион: {}\nСлучаев: {} ({} за сегодня)\nАктивных: {} ({} за сегодня)\n' \
                       'Вылечено: {} ({} за сегодня)\nУмерло: {} ({} за сегодня)'.format(status, region, value[0], plus[0], value[1], plus[1],
                                               value[2], plus[2], value[3], plus[3])

    def _reformat_subject_name(self, name: str or None, week_number: int, ignore_weeks: bool = False) -> str | None:
        """
        Реформат названия предмета с проверкой его присутствия на определённой неделе
        """
        custom_week_pattern = r'кр. ([\d\,]+) н. ([^\\]+)'  # Кроме каких-то недель
        custom_week_range_pattern = r'(\d+\-\d+) н. ([^\\]+)'  # Диапазон
        custom_week_is_set_pattern = r'([\d\,]+) н. ([^\\]+)'  # Включая эти недели
        custom_week_dirt_pattern = r'…'  # Заглушки в расписании
        if name and name != 'None':  # Пара есть?
            data = name.split('\n')
            # Цикл, для сдвоенных пар
            for i in range(len(data)):
                if not ignore_weeks:
                    kr = re.search(custom_week_pattern, data[i])  # Проверяем, есть ли паттерн КР
                    if kr:
                        if str(week_number) in kr.group(1).split(','):  # Если неделя в списке исключённых удаляем
                            data[i] = '--'
                        else:
                            data[i] = kr.group(2)
                    else:
                        range_week = re.search(custom_week_range_pattern, data[i])
                        if range_week:
                            tmp = range_week.group(1).split('-')
                            from_week = int(tmp[0])
                            to_week = int(tmp[1])
                            if from_week <= week_number <= to_week:
                                data[i] = range_week.group(2)
                            else:
                                data[i] = '--'
                        else:
                            is_set = re.search(custom_week_is_set_pattern, data[i])
                            if is_set:
                                if str(week_number) in is_set.group(1).split(','):
                                    data[i] = is_set.group(2)
                                else:
                                    data[i] = '--'
                            else:
                                dirt = re.search(custom_week_dirt_pattern, data[i])
                                if dirt:
                                    data[i] = '--'
            return ' / '.join(data) if data else '--'
        return '--'

    def _reformat_double_pair(self, data: any) -> str:
        """
        Двойные и пустые пары в читабельный формат
        """
        if data:
            if data == 'None':
                return ''
            return ' / '.join(data.split('\n'))
        return '--'

    def _reformat_day_schedule(self, data: list, date: datetime.datetime = datetime.datetime.now(),
                               week_format: bool = False, with_header: bool = True,
                               teacher_header: str | None = None) -> str:
        """
        Форматирует один день из списка в строку для дальнейшего вывода
        """

        result = ''
        if with_header:
            if teacher_header:
                result += '\nРасписание преподавателя {} на {}:\n'.format(
                    teacher_header, self._get_string_date(date, with_week_day=week_format))  # Дата
            else:
                result += '\nРасписание на {}:\n'.format(
                    self._get_string_date(date, with_week_day=week_format))  # Дата
        for i in range(len(data)):
            if len(data[i]) > 1:
                if data[i][0][:len('--')] != '--':
                    result += '{}) {}, {}, {}, {}\n'.format(
                        i + 1,
                        str(data[i][0]),
                        str(data[i][1]) if data[i][1] != '--' and \
                                           data[i][1] != '' else '-',
                        str(data[i][2]) if data[i][2] != '--' and \
                                           data[i][2] != '' else '-',
                        str(data[i][3])) if data[i][3] != '--' and \
                                            data[i][3] != '' else '-'
                else:
                    result += '{}) {}\n'.format(i + 1, '--')
            else:
                result += '{}) {}\n'.format(i + 1, '--')
        return result

    def _reformat_teacher_name(self, teacher: str) -> str:
        """
        Форматирует имя преподавателя
        """
        tmp = teacher.split(' ')
        if len(tmp) == 2:
            teacher = tmp[0].title() + ' ' + tmp[1].upper()
        elif len(tmp) == 1:
            teacher = tmp[0].title()
        return teacher

    def _send_message(self, user_id: int, text: str = '', keyboard: int = 0,
                      custom_keyboard: VkKeyboard = None) -> None:
        """
        Отправка сообщения
        """
        if keyboard == 1:
            keyboard = VkKeyboard(one_time=False)
            keyboard.add_button('на сегодня', color=VkKeyboardColor.POSITIVE)
            keyboard.add_button('на завтра', color=VkKeyboardColor.NEGATIVE)
            keyboard.add_line()
            keyboard.add_button('на эту неделю', color=VkKeyboardColor.PRIMARY)
            keyboard.add_button('на следующую неделю', color=VkKeyboardColor.PRIMARY)
            keyboard.add_line()
            keyboard.add_button('неделя?'.title(), color=VkKeyboardColor.SECONDARY)
            keyboard.add_button('группа?'.title(), color=VkKeyboardColor.SECONDARY)
            keyboard.add_button('помощь'.title(), color=VkKeyboardColor.SECONDARY)
        if custom_keyboard:
            keyboard = custom_keyboard
        try:
            if keyboard:
                self.vk.messages.send(
                    user_id=user_id,
                    random_id=get_random_id(),
                    keyboard=keyboard.get_keyboard(),
                    message=text
                )
            else:
                self.vk.messages.send(
                    user_id=user_id,
                    random_id=get_random_id(),
                    message=text
                )
        except:
            Debug(f'Send message error id: {user_id}', key='ERR')

    def _send_message_with_attachments(self, user_id: int, text: str = '', attachments: list = list) -> None:
        """
        Отправляет сообщение с вложениями
        """
        try:
            self.vk.messages.send(
                user_id=user_id,
                attachment=','.join(attachments),
                random_id=get_random_id(),
                message=text
            )
        except:
            Debug(f'Send message error id: {user_id}', key='ERR')

    def _update_schedule_file(self) -> None:
        """
        Обновляет файл с расписанием
        """
        page = requests.get('https://www.mirea.ru/schedule/')  # Получаем страницу
        soup = BeautifulSoup(page.text, "html.parser")  # Парсим её
        result = soup.find(string="Институт информационных технологий").find_parent("div").find_parent("div").findAll(
            'a', {'class': 'uk-link-toggle'})
        course_pattern = r'([1-3]) курс'
        for x in result:
            course = x.find('div', 'uk-link-heading').text.lower().strip()
            course_number = re.match(course_pattern, course)
            if course_number:
                course_number = course_number.group(1)
                f = open(f'data/schedule{course_number}.xlsx', "wb")
                link = x.get('href')
                resp = requests.get(link)  # запрос по ссылке
                f.write(resp.content)  # Записываем контент в файл
                f.close()
        self.last_schedule_file_update = time.time()
        Debug('Update schedule files', key='UPD')
        self._parse_schedule_file()

    def _parse_schedule_file(self) -> None:
        """
        Парсит полученные файлы расписание и записывает в списки
        """
        self.schedule_data = []
        for c in range(3):
            book = openpyxl.load_workbook(
                f'data/schedule{c + 1}.xlsx')  # открытие файла
            sheet = book.active  # активный лист
            num_cols = sheet.max_column  # количество столбцов
            last_group_cell = 0  # Сколько прошло ячеек от последней группы
            for i in range(6, num_cols):
                if last_group_cell >= 4:  # Если после группы прошло 4 ячейки, ждём следующей группы
                    last_group_cell = -1
                    continue
                column = []
                for j in range(2, 76):  # Перебираем
                    v = str(sheet.cell(column=i, row=j).value)
                    if j == 2 and re.match(r'\w{4}-\d{2}-\d{2}',
                                           v):  # Если ячейка вторая, то проверяем что это номер группы
                        last_group_cell = 0  # Если это так, обнуляем счётчик
                    column.append(v)
                if last_group_cell != -1:  # Пока не дошли до следующей группы, не добавляем столбцы,
                    self.schedule_data.append(column)
                    last_group_cell += 1
        Debug('Parse schedule file', key='PRS')


def main():
    bot = VkBot()
    bot.start_listen()


if __name__ == '__main__':
    while True:
        main()
